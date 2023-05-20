import json
from openpyxl import load_workbook
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.views import View
from itasset.my_forms import *
from django.db.models import Q,Sum,Count
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.mixins import LoginRequiredMixin
from django.utils.decorators import method_decorator

# Search view
class SearchView(LoginRequiredMixin,View):
    # login_url = '/accounts/login/'
    def get(self,request):
        # Get asset list
        group_list = [obj.name for obj in request.user.groups.all()]
        asset_list = Asset.objects.filter(company__in=group_list)
        user_list = User.objects.filter(company__in=group_list)
        # Filter assets
        if request.GET.get('keywords'):
            keywords = request.GET.get('keywords')
            asset_list = asset_list.filter(Q(name__icontains=keywords)|Q(sn__icontains=keywords)|Q(user__name__icontains=keywords)|Q(user__account__icontains=keywords))
            # Get user list
            user_list = user_list.filter(Q(name__icontains=request.GET.get('keywords'))|Q(account__icontains=request.GET.get('keywords')))
        return render(request,'itasset/search.html',{'keywords':request.GET.get('keywords'),'asset_list':asset_list,'user_list':user_list})


# Home view
class Home(LoginRequiredMixin,View):
    def get(self,request):
        group_list = [obj.name for obj in request.user.groups.all()]
        if not request.GET:
            return render(request,'itasset/home.html',locals())        
        asset_list = Asset.objects.filter(company__in=group_list)        
        # Get summary count
        annotate_data = list(asset_list.values('company','purchase_date__year','category').annotate(count=Count('company'),total=Sum('price')))     
        # Get company list
        company_list = list(set([item['company'] for item in annotate_data]))
        company_list.sort()
        # Create variable to store structure data
        data = {}
        for company in company_list:
            data[company] = {}
            # Get category list for each company
            category_list = list({item['category'] for item in annotate_data if company in item.get('company')})
            # Remove empty and sort
            if '' in category_list:
                category_list.remove('')
            category_list.sort()
            # Get year list for each company
            year_list = list({item['purchase_date__year'] for item in annotate_data if company in item.get('company')})
            # Remove empty and sort
            if '' in year_list:
                year_list.remove('')
            year_list.sort()
            if len(year_list) == 1:
                year_list = [year_list[0],year_list[0]-1]
            data[company] = [year_list,{}]
            # Get count for each category in each year
            for category in category_list:
                data[company][1][category] = [[],[]]
                for year in year_list:
                    current_count = 0
                    current_price = 0
                    for item in annotate_data:
                        if company == item.get('company') and category == item.get('category') and year == item.get('purchase_date__year'):
                            current_count = item['count']
                            current_price = item['total']
                    data[company][1][category][0].append(current_count)
                    data[company][1][category][1].append(int(current_price))

            # Get tree summary
            tree_summary = {'name':'Assets','children':[]}
            annotate_summary = list(asset_list.values('company','category','status__name').annotate(count=Count('company'))) 
            i = 0
            for company in company_list:
                tree_summary['children'].append({'name':company,'children':[]})
                category_list = list({item['category'] for item in annotate_summary if company in item.get('company')})
                i2 = 0
                for category in category_list:
                    tree_summary['children'][i]['children'].append({'name':category,'children':[]})
                    status_counts = [
                        {'name':'%s  (%s)'%(item['status__name'],item['count']),'value':item['count']} 
                        for item in annotate_summary if company == item.get('company') and category == item.get('category')
                        ]
                    tree_summary['children'][i]['children'][i2]['children'] = (status_counts)
                    i2 += 1
                i += 1
      

        return HttpResponse(json.dumps({'status':True,'data':data,'tree_summary':tree_summary}))


# View for asset list, add, clone, edit, approve, assign, turnover, scrap and scrapapprove
class AssetView(LoginRequiredMixin,View):

    def get(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','add','clone','edit','approve','assign','turnover','scrap','scrapapprove']:
            # Get asset and user list
            group_list = [obj.name for obj in request.user.groups.all()]
            asset_list = Asset.objects.filter(company__in=group_list).order_by('-id')   
            user_list = User.objects.filter(company__in=group_list).order_by('-id')   
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the asset to modify exist
        if action in ['detail','clone','edit','approve','assign','turnover','scrap','scrapapprove']:
            if request.GET.get('pk'):
                asset = asset_list.filter(pk=request.GET.get('pk'))
                if asset.exists():
                    asset = asset.first()
                    if asset.company not in group_list:
                        return HttpResponse('Permission denied!')
                else:
                    return HttpResponse('The asset do not exist!')
            else:
                return HttpResponse('Please provide pk for the asset!')                    
             
        # Return list page
        if action == 'list':           
            form = AssetFilterForm() 
            annotate_data = list(asset_list.values('company','category').annotate(count=Count('company'),total=Sum('price')))  
            # Filter data and create filter form
            if request.GET:
                form = AssetFilterForm(request.GET)
                for k,v in dict(request.GET).items():
                    if v[0] != '':
                        asset_list = asset_list.filter(**{k:v[0]})

            # Change filter form choices               
            # Get company list
            company_list = list(set([item['company'] for item in annotate_data if item['company'] is not None and item['company'] != '' and item['company'] != 'None']))
            company_list.sort()
            form.fields['company'].choices = [('','All Company')] + [(item,item) for item in company_list] 
            # Get category list
            category_list = list(set([item['category'] for item in annotate_data if item['category'] is not None and item['category'] != '' and item['category'] != 'None']))
            category_list.sort()
            form.fields['category'].choices = [('','All Category')] + [(item,item) for item in category_list]  
            return render(request,'itasset/asset/list.html',locals())
        
         # Return page for add, clone and edit
        if action in ['add','clone','edit']:
            title = 'Asset ' + action
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            if action == 'add':
                form = AssetForm()

            if action in ['clone','edit']:
                form = AssetForm(instance=asset)    

            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name').order_by('name'))

            return render(request,'itasset/asset/form.html',locals())    

        # Return detail page
        if action == 'detail':
            user_list = user_list.filter(is_active=True)
            history = AssetHistory.objects.filter(asset=asset).order_by('-id')
            return render(request,'itasset/asset/detail.html',locals())

        # Return approve page
        if action== 'approve':
            if 'Approver' not in group_list:
                return HttpResponse('Permission denied!')
            return render(request,'itasset/asset/approve.html',{'asset':asset})
            

        # Return assign page
        if action == 'assign':
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            user_to_assign = User.objects.filter(pk=request.GET.get('uid')).first()
            return render(request,'itasset/asset/assign.html',{'asset':asset,'user_to_assign':user_to_assign})

        # Return turnover page
        if action == 'turnover':
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            return render(request,'itasset/asset/turnover.html',{'asset':asset})

        # Return scrap page
        if action == 'scrap':
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            return render(request,'itasset/asset/scrap.html',{'asset':asset})

        # Return scrapapprove page
        if action == 'scrapapprove':
            if 'Approver' not in group_list:
                return HttpResponse('Permission denied!')
            return render(request,'itasset/asset/scrapapprove.html',{'asset':asset})

    def post(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','add','clone','edit','approve','assign','turnover','scrap','scrapapprove']:
            # Get asset and user list
            group_list = [obj.name for obj in request.user.groups.all()]
            asset_list = Asset.objects.filter(company__in=group_list).order_by('-id')      
            user_list = User.objects.filter(company__in=group_list).order_by('-id')   
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the asset to modify exist
        if action in ['clone','edit','approve','assign','turnover','scrap','scrapapprove']:
            if request.GET.get('pk'):
                asset = Asset.objects.filter(pk=request.GET.get('pk'))
                if asset.exists():
                    asset = asset.first()
                else:
                    return HttpResponse('The asset do not exist!')
            else:
                return HttpResponse('Please provide pk for the asset!')   

        # Run add or clone logic
        if action in ['add','clone']:
            title = 'Asset ' + action
            form = AssetForm(request.POST)
            if form.is_valid():
                form.instance.create_by = request.user.username
                new_asset = form.save()
                # Add to history
                AssetHistory.objects.create(asset=new_asset,action="Create",operator=request.user.username,comment=f'Create {new_asset.name}')
                return redirect('/itasset/asset/detail' + '?pk=%s'%(new_asset.id)) 
            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name'))     
            return render(request,'itasset/asset/form.html',{'form':form,'title':title})

        # Run edit logic
        if action == 'edit':
            title = 'Asset ' + action  
            asset_before_edit = dict(asset.__dict__)       
            form = AssetForm(request.POST,instance=asset)
            if form.is_valid():
                if form.changed_data:       
                    form.save()
                    # Check what changed
                    comment = []
                    for item in form.changed_data:
                        comment.append('Change %s from %s to %s'%(item,asset_before_edit[item],form.cleaned_data[item]))
                    # Add to history
                    AssetHistory.objects.create(asset=asset,action="Edit",operator=request.user.username,comment=', '.join(comment))
                return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id)) 
            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name'))              
            return render(request,'itasset/asset/form',{'form':form,'title':title})
        
        # Run approve logic
        if action== 'approve':
            #Only available asset can assign
            if asset.status_id == 1:
                try:
                    #Approve asset                   
                    asset.status_id = 2
                    asset.save()
                    #Add to history
                    comment = 'Approve asset %s'%(asset.name)
                    AssetHistory.objects.create(asset=asset,action="Approve",operator=request.user.username,comment=comment)
                    return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id))                    
                except Exception as e:
                    print(e)
            msg =  f'Can only approve pending approve asset, current asset status is {asset.status}!'             
            return render(request,'itasset/asset/approve.html',{'asset':asset,'msg':msg})
        
        # Run assign logic
        if action == 'assign':
            #Get employee
            user_to_assign = User.objects.filter(pk=request.GET.get('uid')).first()
            #Can only assign available asset to active user
            if asset.status_id == 2 and user_to_assign.is_active:
                try:
                    #Assign asset to employee                    
                    asset.user = user_to_assign
                    asset.status_id = 3
                    asset.save()
                    #Add to history
                    comment = 'Assign asset %s to %s'%(asset.name,user_to_assign.name)
                    AssetHistory.objects.create(asset=asset,action="Assign",operator=request.user.username,comment=comment)
                    return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id))                    
                except Exception as e:
                    print(e)
            msg =  f'Can only assign available asset to active user, current asset status is {asset.status}, user status is {user_to_assign.is_active}!'             
            return render(request,'itasset/asset/assign.html',{'asset':asset,'user_to_assign':user_to_assign,'msg':msg})

        # Run turnover logic
        if action == 'turnover':
            #Can only turn over assigned asset
            if asset.status_id == 3:
                try:
                    #Approve asset                   
                    asset.status_id = 2
                    asset.user = None
                    asset.save()
                    #Add to history
                    comment = 'Turn over asset %s'%(asset.name)
                    AssetHistory.objects.create(asset=asset,action="Turn Over",operator=request.user.username,comment=comment)
                    return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id))                    
                except Exception as e:
                    print(e)
            msg =  f'Can only turn over assigned asset, current asset status is {asset.status}!'             
            return render(request,'itasset/asset/turnover.html',{'item':asset,'msg':msg})

        # Run scrap logic
        if action == 'scrap':
            #Can only scrap available asset
            if asset.status_id == 2:
                try:
                    #Approve asset                   
                    asset.status_id = 4
                    asset.save()
                    #Add to history
                    comment = 'Scrap asset %s'%(asset.name)
                    AssetHistory.objects.create(asset=asset,action="Scrap",operator=request.user.username,comment=comment)
                    return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id))                    
                except Exception as e:
                    print(e)
            msg =  f'Can only scrap available asset, current asset status is {asset.status}!'             
            return render(request,'itasset/asset/scrap.html',{'item':asset,'msg':msg})

        # Run scrapapprove logic
        if action == 'scrapapprove':
            #Can only scrap approve scrap request asset
            if asset.status_id == 4:
                try:
                    #Approve asset                   
                    asset.status_id = 5
                    asset.save()
                    #Add to history
                    comment = 'Scrap approve asset %s'%(asset.name)
                    AssetHistory.objects.create(asset=asset,action="Scrap Approve",operator=request.user.username,comment=comment)
                    return redirect('/itasset/asset/detail' + '?pk=%s'%(asset.id))                  
                except Exception as e:
                    print(e)
            msg =  f'Can only scrap approve scrap request asset, current asset status is {asset.status}!'             
            return render(request,'itasset/asset/scrapapprove.html',{'item':asset,'msg':msg})


# View for user list, add, clone, resign and edit
class UserView(LoginRequiredMixin,View):

    def get(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','add','clone','edit','resign']:
            # Get asset and user list
            group_list = [obj.name for obj in request.user.groups.all()]
            asset_list = Asset.objects.filter(company__in=group_list).order_by('-id')      
            user_list = User.objects.filter(company__in=group_list).order_by('-id')   
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the user to modify exist
        if action in ['detail','clone','edit','resign']:
            if request.GET.get('pk'):
                user = user_list.filter(pk=request.GET.get('pk'))
                if user.exists():
                    user = user.first()
                else:
                    return HttpResponse('The user do not exist!')
            else:
                return HttpResponse('Please provide pk for the user!')                    
             
        # Return list page
        if action == 'list':           
            form = UserFilterForm() 
            annotate_data = list(asset_list.values('company','category').annotate(count=Count('company'),total=Sum('price')))  
            # Filter data and create filter form
            if request.GET:
                form = UserFilterForm(request.GET)
                for k,v in dict(request.GET).items():
                    if v[0] != '':
                        user_list = user_list.filter(**{k:v[0]}) 

            # Change filter form choices                
            # Get company list
            company_list = list(set([item['company'] for item in annotate_data]))
            while '' in company_list:
                company_list.remove('') 
            company_list.sort()
            form.fields['company'].choices = [('','All Company')] + [(item,item) for item in company_list]  
            return render(request,'itasset/user/list.html',locals())
        
         # Return page for add, clone and edit
        if action in ['add','clone','edit']:
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            title = 'User ' + action 
            if action == 'add':
                form = UserForm()

            if action in ['clone','edit']:
                form = UserForm(instance=user)    
            
            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name').order_by('name'))
            return render(request,'itasset/user/form.html',locals())    

        # Return detail page
        if action == 'detail':
            available_asset_list = asset_list.filter(status_id=2)
            asset_list = Asset.objects.filter(user=user)
            history = UserHistory.objects.filter(user=user).order_by('-id')
            return render(request,'itasset/user/detail.html',locals())

        
         # Return resign page
        if action == 'resign':
            if 'Editor' not in group_list:
                return HttpResponse('Permission denied!')
            if request.GET.get('pk'):
                asset_list = Asset.objects.filter(user=user)
                return render(request,'itasset/user/resign.html',locals())
        

    def post(self,request,action):
        # Check if action in defined list
        if action in ['add','clone','edit','resign']:
            # Get asset and user list
            group_list = [obj.name for obj in request.user.groups.all()]
            asset_list = Asset.objects.filter(company__in=group_list).order_by('-id')      
            user_list = User.objects.filter(company__in=group_list).order_by('-id')   
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the user to modify exist
        if action in ['clone','edit','resign']:
            if request.GET.get('pk'):
                user = user_list.filter(pk=request.GET.get('pk'))
                if user.exists():
                    user = user.first()
                else:
                    return HttpResponse('The user do not exist!')
            else:
                return HttpResponse('Please provide pk for the user!')    

        # Run add or clone logic
        if action in ['add','clone']:
            title = 'User ' + action 
            form = UserForm(request.POST)
            if form.is_valid():
                form.instance.create_by = request.user.username
                new_user = form.save()
                # Add to history
                UserHistory.objects.create(user=new_user,action="Create",operator=request.user.username,comment=f'Create {new_user.name}')
                return redirect('/itasset/user/detail' + '?pk=%s'%(new_user.id))  
            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name'))    
            return render(request,'itasset/user/form.html',{'form':form,'title':title})

        # Run edit logic
        if action == 'edit':
            title = 'User' + action 
            user_before_edit = dict(user.__dict__)           
            form = UserForm(request.POST,instance=user)
            if form.is_valid():
                if form.changed_data:       
                    form.save()
                    # Check what changed
                    comment = []
                    for item in form.changed_data:
                        comment.append('Change %s from %s to %s'%(item,user_before_edit[item],form.cleaned_data[item]))
                    # Add to history
                    UserHistory.objects.create(user=user,action="Edit",operator=request.user.username,comment=', '.join(comment))
                return redirect('/itasset/user/detail' + '?pk=%s'%(user.id))   
            # Change company choices
            form.fields['company'].choices = list(Company.objects.filter(name__in=group_list).values_list('name','name'))              
            return render(request,'itasset/user/form',{'form':form,'title':title})
        
        # Run resign logic
        if action == 'resign':
            user = User.objects.filter(pk=request.GET.get('pk'))
            asset_list = Asset.objects.filter(user=user.first())                    
            if asset_list.exists():
                # Add history
                for obj in asset_list:
                    AssetHistory.objects.create(asset=obj,action='Turn Over',operator=request.user.username,comment='Turn over when resign')
                # Turn over assets
                asset_list.update(status=2,user=None)
            # Resign user        
            user.update(is_active=False) 
            # Add history
            UserHistory.objects.create(user=user.first(),action='Resign',operator=request.user.username)           
            return redirect('/itasset/user/detail?pk=%s'%(user.first().id))
        

# View to import asset and user
class ImportView(LoginRequiredMixin,View):     

    def get(self,request,tablename):
        # Check if the table is defined
        if tablename not in ['asset','user','approve','assign','scrap','scrapapprove']:
            return HttpResponse('The table name is not defined!')
        if tablename == 'asset':
            title = 'Asset Import'
        if tablename == 'user':
            title = 'User Import'
        if tablename == 'approve':
            title = 'Bulk approve'
        if tablename == 'assign':
            title = 'Bulk assign'
        if tablename == 'scrap':
            title = 'Bulk scrap'
        if tablename == 'scrapapprove':
            title = 'Bulk scrap approve'
        return render(request,'itasset/import/data_import.html',locals())
        pass
    def post(self,request,tablename):
        # Check if the table is defined
        if tablename not in ['asset','user','approve','assign','scrap','scrapapprove']:
            return HttpResponse('The table name is not defined!')

        # Get excel file
        excel_file = request.FILES['csv_file']
        # Get file data
        wb = load_workbook(excel_file,read_only=True,data_only=True)
        sheet_obj = wb[wb.sheetnames[0]]
        # Get table head
        current_row = 0
        result_list = []
        for row in sheet_obj.iter_rows(values_only=True):
            if current_row == 0:
                current_row += 1
                tableh = list(row)
                continue
            else:
                row = list(row)
                # Convert data to dictionary
                data_dic = {}
                i = 0
                for column_name in tableh:
                    data_dic[column_name] = row[i]
                    i += 1
            current_row += 1

            # Import assets:
            if tablename == 'asset':
                title = 'Result for asset import'
                try:
                    new_asset = Asset(**data_dic)
                    new_asset.create_by = request.user.username
                    if data_dic.get('create_by','') != '':
                        new_asset.create_by = data_dic.get('create_by')                                      
                    new_asset.save()  
                    if data_dic.get('create_time','') != '':
                        new_asset.create_time = data_dic.get('create_time')  
                        new_asset.save()                                       
                    # Add history
                    AssetHistory.objects.create(asset=new_asset,action='Create',operator=request.user.username,comment='Bulk Import')
                    result = 'Success'
                except Exception as e:
                    result =  str(e)

            # Import users:
            if tablename == 'user':
                title = 'Result for asset import'
                form = UserForm(data=data_dic)
                try:
                    new_user = User(**data_dic)
                    new_user.create_by = request.user.username
                    new_user.save()                    
                    # Add history
                    UserHistory.objects.create(user=new_user,action='Create',operator=request.user.username,comment='Bulk Import')
                    result = 'Success'
                except Exception as e:
                    result =  str(e)

            # Approve asset:
            if tablename == 'approve':
                title = 'Result for bulk approve'
                try:
                    asset_to_approve = Asset.objects.filter(sn=data_dic.get('sn'))
                    if asset_to_approve.first().status_id == 1:
                        asset_to_approve.update(status=2)                   
                        # Add history
                        AssetHistory.objects.create(asset=asset_to_approve.first(),action='Approve',operator=request.user.username,comment='Bulk Import')
                        result = 'Success'
                    else:
                        result = 'Failed, only pending approve asset can approve, current asset status is %s'%(asset_to_approve.first().status)
                except Exception as e:
                    result =  str(e)

            # Assign asset:
            if tablename == 'assign':
                title = 'Result for bulk assign'
                try:
                    asset_to_assign = Asset.objects.filter(sn=data_dic.get('sn'))
                    if asset_to_assign.first().status_id == 2:
                        user_to_assign = User.objects.filter(account=data_dic.get('account'))
                        if user_to_assign.exists():
                            user_to_assign = user_to_assign.first()
                            asset_to_assign.update(user=user_to_assign,status=3)                   
                            # Add history
                            AssetHistory.objects.create(asset=asset_to_assign.first(),action='Assign',operator=request.user.username,comment='Bulk Import')
                            result = 'Success'
                        else:
                            result = 'Failed, the user account do no exist'                                      
                    else:
                        result = 'Failed, only available asset can assign, current asset status is %s'%(asset_to_assign.first().status)
                except Exception as e:
                    result =  str(e)

            # Assign asset:
            if tablename == 'scrap':
                title = 'Result for bulk scrap'
                try:
                    asset_to_scrap = Asset.objects.filter(sn=data_dic.get('sn'))
                    if asset_to_scrap.first().status_id == 2:
                        asset_to_scrap = User.objects.filter(account=row[1]).first()
                        asset_to_scrap.update(user=user_to_assign,status=4)                   
                        # Add history
                        AssetHistory.objects.create(asset=asset_to_scrap.first(),action='Scrap',operator=request.user.username,comment='Bulk Import')
                        result = 'Success'
                    else:
                        result = 'Failed, only available asset can scrap, current asset status is %s'%(asset_to_scrap.first().status)
                except Exception as e:
                    result =  str(e)

            # Assign asset:
            if tablename == 'scrapapprove':
                title = 'Result for bulk scrap approve'
                try:
                    asset_to_scrap = Asset.objects.filter(sn=data_dic.get('sn'))
                    if asset_to_scrap.first().status_id == 4:
                        asset_to_scrap.update(user=user_to_assign,status=5)                   
                        # Add history
                        AssetHistory.objects.create(asset=asset_to_scrap.first(),action='Scrap approve',operator=request.user.username,comment='Bulk Import')
                        result = 'Success'
                    else:
                        result = 'Failed, only scrap request asset can scrap approve, current asset status is %s'%(asset_to_scrap.first().status)
                except Exception as e:
                    result =  str(e)

            result_list.append([result] + row)    
        tableh = ['Result'] + tableh                   
        return render(request,'itasset/import/import_result.html',locals())

    
# View for reports
class ReportView(LoginRequiredMixin,View):
    def get(self,request,reportname):
        # Check if the report name exist
        if reportname in ['custom_report','tanium']:
            group_list = [obj.name for obj in request.user.groups.all()]
            asset_list = Asset.objects.filter(company__in=group_list) 
        else:
            return HttpResponse('The report name do not exist!')
        
        # Generate custom report
        if reportname == 'custom_report':
            form = AssetCustomFilterForm()
            annotate_data = list(asset_list.values('company','category','office','os').annotate(count=Count('company'),total=Sum('price'))) 
            # Filter data and create filter form
            if request.GET:
                form = AssetCustomFilterForm(request.GET)
                for k,v in dict(request.GET).items():
                    if v[0] != '':
                        asset_list = asset_list.filter(**{k:v[0]})

            # Change filter form choices             
            # Get company list
            company_list = list(set([item['company'] for item in annotate_data if item['company'] is not None and item['company'] != '' and item['company'] != 'None']))
            company_list.sort()
            form.fields['company'].choices = [('','All Company')] + [(item,item) for item in company_list] 
            # Get category list
            category_list = list(set([item['category'] for item in annotate_data if item['category'] is not None and item['category'] != '' and item['category'] != 'None']))
            category_list.sort()
            form.fields['category'].choices = [('','All Category')] + [(item,item) for item in category_list] 
            # Get office list
            office_list = list(set([item['office'] for item in annotate_data if item['office'] is not None and item['office'] != '' and item['office'] != 'None']))
            office_list.sort()
            form.fields['office'].choices = [('','All Office')] + [(item,item) for item in office_list] 
            # Get os list
            os_list = list(set([item['os'] for item in annotate_data if item['os'] is not None and item['os'] != '' and item['os'] != 'None']))
            # os_list.sort()
            form.fields['os'].choices = [('','All OS')] + [(item,item) for item in os_list] 
                    
            return render(request,'itasset/report/custom_report.html',locals())

        # Generate tanium report
        if reportname == 'tanium':
            tanium_list = Tanium.objects.all()
            return render(request,'itasset/report/tanium.html',{'tanium_list':tanium_list})


# View for API
@method_decorator(csrf_exempt, name='dispatch')
class API(View):
    def post(self,request,apiname):
        # Check if apiname exist
        if apiname not in ['tanium']:
            return HttpResponse(json.dumps({'status':False,'error':'The api do not exist!'}))
        if apiname == 'tanium':
            # Check if the record exist
            check_exist = Tanium.objects.filter(name=request.POST.get('name'))
            if check_exist.exists():
                form = TaniumForm(data=request.POST,instance=check_exist.first())
            else:
                form = TaniumForm(data=request.POST)
            if form.is_valid():
                try:
                    form.save()
                    result = {'status':True}
                except Exception as e:
                    result = {'status':False,'error':str(e)}
            else:
                result = {'status':False,'error':str(form.errors)}
            return HttpResponse(json.dumps(result))

