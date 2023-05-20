from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views import View
from itdc.models import *
from itdc.my_forms import *
import json
from itdc.my_functions import  draw_diagram
from django.db.models import Q,Sum,Count
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.mixins import LoginRequiredMixin,PermissionRequiredMixin
from django.utils.decorators import method_decorator
from openpyxl import load_workbook

# Search view
class SearchView(LoginRequiredMixin,PermissionRequiredMixin,View):
    permission_required = 'itdc.view_asset'
    def get(self,request):
        # Get keywords
        keywords = request.GET.get('keywords')
        # Filter assets and software
        if request.GET.get('keywords'):
            asset_list = Asset.objects.filter(name__icontains=keywords)
            software_list = Software.objects.filter(name__icontains=keywords)
        return render(request,'itdc/search.html',locals())

# View for home
class HomeView(LoginRequiredMixin,PermissionRequiredMixin,View):
    permission_required = 'itdc.view_asset'
    def get(self,request):
        if not request.GET:
            return render(request,'itdc/home.html')
        asset_summary = list(Asset.objects.filter(is_active=True,is_poweron=True).values('category').annotate(value=Count('category')).order_by('-value'))
        asset_summary = [{'name':item['category'],'value':item['value']} for item in asset_summary]
        os_summary = list(Asset.objects.filter(is_active=True,is_poweron=True).values('os').annotate(value=Count('os')).order_by('-value'))
        os_summary = [{'name':item['os'],'value':item['value']} for item in os_summary if item['os']]
        location_summary = list(Asset.objects.filter(is_active=True,is_poweron=True).values('location').annotate(value=Count('location')).order_by('-value'))
        location_summary = [{'name':item['location'],'value':item['value']} for item in location_summary]
        software_summary = list(Software.objects.values('name').annotate(value=Count('name')).order_by('-value'))

        # Get tree summary
        annotate_data = list(Asset.objects.filter(is_active=True,is_poweron=True).values('location','category').annotate(count=Count('location')))     
        # Get location list
        location_list = list(set([item['location'] for item in annotate_data]))
        tree_data = {'name':'Assets','children':[]}
        i = 0
        for location in location_list:
            tree_data['children'].append({'name':location,'children':[]})
            for item in annotate_data:
                if location == item['location']:
                    tree_data['children'][i]['children'].append({'name':'%s  (%s)'%(item['category'],item['count']),'value':item['count']})
            i += 1

        return HttpResponse(json.dumps({'status':True,'data':[asset_summary,os_summary,location_summary,software_summary,tree_data]}))
       
# View for asset
@method_decorator(csrf_exempt, name='dispatch')
class AssetView(LoginRequiredMixin,PermissionRequiredMixin,View):  
    permission_required = 'itdc.view_asset'  
    model_class = Asset
    # form_class = PMForm
    list_template = 'itdc/asset/list.html'
    form_template = 'itdc/asset/form.html'
    detail_template = 'itdc/asset/detail.html'
    delete_template = 'itdc/asset/delete.html'
    list_url = '/itdc/asset/list'
    def get_form_class(self,category):
        if category == 'PM':
            self.form_class = PMForm
        elif category == 'VM':
            self.form_class = VMForm
        elif category == 'Storage':
            self.form_class = StorageForm
        elif category == 'Cluster':
            self.form_class = ClusterForm
        elif category == 'Switch':
            self.form_class = SwitchForm
        else:
            self.form_class = OtherForm


    def get(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','add','clone','edit','delete']:
            # Get asset list
            asset_list = Asset.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the asset exist
        if action in ['detail','clone','edit','delete']:
            if request.GET.get('pk'):
                asset = asset_list.filter(pk=request.GET.get('pk'))
                if asset.exists():
                    asset = asset.first()
                else:
                    return HttpResponse('The asset do not exist!')
            else:
                return HttpResponse('Please provide pk for the asset!')  

        # Get form class
        category = request.GET.get('category')
        self.get_form_class(category=category)                  
             
        # Return list page
        if action == 'list':
            form = AssetFilterForm() 
            if request.GET:
                form = AssetFilterForm(request.GET) 
                for k,v in dict(request.GET).items():
                    if v[0] != '':
                        asset_list = asset_list.filter(**{k:v[0]})                   
            return render(request,self.list_template,locals())
        
         # Return page for add, clone and edit
        if action in ['add','clone','edit']:
            title = category + ' ' + action
            if action == 'add':
                form = self.form_class()

            if action in ['clone','edit']:
                form = self.form_class(instance=asset)  
            
            return render(request,self.form_template,locals())    

        # Return detail page
        if action == 'detail':
            portaddform = PortAddForm()   
            portform = PortForm() 
            commentaddform = AssetCommentForm()  
            # Check if cotained asset exist
            if asset.is_container:
                contained_assets = Asset.objects.filter(container=asset)
            return render(request,self.detail_template,locals())   

        # Return delete page
        if action == 'delete':            
            return render(request,self.delete_template,{'list_url':self.list_url,'asset':asset})  
              

    def post(self,request,action):
        # Check if action in defined list
        if action in ['add','clone','edit','delete','addcomment','deletecomment']:
            # Get asset list
            asset_list = Asset.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the asset exist
        if action in ['clone','edit','delete','addcomment']:
            if request.GET.get('pk'):
                asset = asset_list.filter(pk=request.GET.get('pk'))
                if asset.exists():
                    asset = asset.first()
                else:
                    return HttpResponse('The asset do not exist!')
            else:
                return HttpResponse('Please provide pk for the asset!')      

        # Get form class
        if request.GET.get('category'):
            category = request.GET.get('category')
            self.get_form_class(category=category)
            title = category + ' ' + action

        # Run add or clone logic
        if action in ['add','clone']:            
            form = self.form_class(request.POST)
            if form.is_valid():
                form.instance.create_by = request.user.username
                form.instance.category = category
                form.save()
                return redirect(self.list_url)  
            return render(request,self.form_template,{'form':form,'title':title})

        # Run edit logic
        if action == 'edit':       
            form = self.form_class(request.POST,instance=asset)
            if form.is_valid():
                if form.changed_data:       
                    form.save()
                return redirect(self.list_url)              
            return render(request,self.form_template,{'form':form,'title':title})


        # Run delete logic
        if action == 'delete':
            asset_list.filter(pk=request.GET.get('pk')).delete()
            return redirect(self.list_url) 

        # Run add comment logic
        if action == 'addcomment':
            form = AssetCommentForm(request.POST)
            if form.is_valid():
                try:
                    AssetComment.objects.create(asset=asset,comment=request.POST.get('comment'),create_by=request.user.username)  
                    result = {'status':True,}
                except Exception as e:
                    print(e)
                    result = {'status':False,'msg':'Failed to add comment!'}
            else:
                result = {'status':False,'msg':'Please add correct comment!'}

            return HttpResponse(json.dumps(result))

        # Run delete comment logic
        if action == 'deletecomment':
            if request.POST.get('commentid'):
                try:
                    AssetComment.objects.filter(pk=request.POST.get('commentid')).delete()
                    result = {'status':True,}
                except Exception as e:
                    print(e)
                    result = {'status':False,'msg':'Failed to delete comment!'}
            else:
                result = {'status':False,'msg':'Please provide corret ID for comment!'}

            return HttpResponse(json.dumps(result))

# View for network ports
@method_decorator(csrf_exempt, name='dispatch')
class PortView(LoginRequiredMixin,View):
    def get(self,request,action):
        # Get network port detail
        if action == 'detail':
            network_port = NetworkPort.objects.filter(pk=request.GET.get('pk'))
            if network_port.exists():
                result = {'status':True,'data':network_port.values('port_number','connect_to','comment').first()}
            else:
                result = {'status':False,'error':'The port do not exist!'}
            return HttpResponse(json.dumps(result))
    def post(self,request,action):

        # Check if the action exist
        if action not in ['add','edit','delete']:
            return HttpResponse(json.dumps({'status':False,'msg':'The action do not exist'}))

        # Check if the port exist
        if action in ['edit','delete']:
            network_port = NetworkPort.objects.filter(pk=request.GET.get('pk'))
            if not network_port.exists():
                return HttpResponse(json.dumps({'status':False,'msg':'The network port do not exist'}))  

        # Run add logic
        if action == 'add':
            form = PortAddForm(request.POST)
            if form.is_valid():
                current_port = int(request.POST.get('port_start')) 
                port_end = int(request.POST.get('port_end')) 
                port_list = []
                while current_port <= port_end:
                    port_list.append(NetworkPort(port_number=current_port,asset_id=request.GET.get('assetid'),create_by=request.user.username))
                    current_port = current_port + 1
                try:
                    NetworkPort.objects.bulk_create(port_list)
                    result = {'status':True,'msg':'Port(s) added!'}
                except Exception as e:
                    result = {'status':False,'msg':'Port(s) failed to add!'}
                    print(e)
            else:
                result = {'status':False,'msg':'Please type in correct port information!'}
            return HttpResponse(json.dumps(result))

        # Run edit logic
        if action == 'edit':
            form = PortForm(request.POST,instance=network_port.first())
            if form.is_valid():
                form.save()
                result = {'status':True,'msg':'Port edited!'}
            else:
                result = {'status':False,'msg':'Failed to edit!'}
            return HttpResponse(json.dumps(result))

        # Run delete logic
        if action == 'delete':
            try:
                network_port.delete()
                result = {'status':True,'msg':'Port was deleted!'}
            except Exception as e:
                result = {'status':False,'msg':'Failed to delete!'}
            return HttpResponse(json.dumps(result))

# View for software
class SoftwareView(LoginRequiredMixin,PermissionRequiredMixin,View): 
    permission_required = 'itdc.view_software'  
    model_class = Software
    form_class = SoftwareForm
    list_template = 'itdc/software/list.html'
    form_template = 'itdc/software/form.html'
    detail_template = 'itdc/software/detail.html'
    delete_template = 'itdc/software/delete.html'
    list_url = '/itdc/software/list'
    def get(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','add','clone','edit','delete','assign','unassign']:
            # Get software list
            software_list = Software.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the software exist
        if action in ['detail','clone','edit','delete','assign','unassign']:
            if request.GET.get('pk'):
                software = software_list.filter(pk=request.GET.get('pk'))
                if software.exists():
                    software = software.first()
                else:
                    return HttpResponse('The software do not exist!')
            else:
                return HttpResponse('Please provide pk for the software!')                    
             
        # Return list page
        if action == 'list':                     
            return render(request,self.list_template,locals())
        
         # Return page for add, clone and edit
        if action in ['add','clone','edit']:
            if action == 'add':
                title = 'Software Add'
                form = self.form_class()

            if action == 'clone':
                title = 'Software Clone'
                form = self.form_class(instance=software)    

            if action == 'edit':
                title = 'Software Edit'
                form = self.form_class(instance=software)   
            
            return render(request,self.form_template,locals())    

        # Return detail page
        if action == 'detail':  
            available_assets = Asset.objects.filter(category__in=['pm','vm'],is_active=True).order_by('-id')          
            return render(request,self.detail_template,locals())   

        # Return delete page
        if action == 'delete':            
            return render(request,self.delete_template,{'list_url':self.list_url,'software':software})  

        # Run assign logic
        if action == 'assign':
            asset = Asset.objects.filter(pk=request.GET.get('assetid')).first()
            try:
                software.asset.add(asset)
                status = {'status':True,'msg':'Finish assign successfully!'}
            except Exception as e:
                status = {'status':False,'msg':'Failed to assign, please try again!'}
                print(e)
            return HttpResponse(json.dumps(status))

        # Run unassign logic
        if action == 'unassign':
            asset = Asset.objects.filter(pk=request.GET.get('assetid')).first()
            try:
                software.asset.remove(asset)
                status = {'status':True,'msg':'Finish unassign successfully!'}
            except Exception as e:
                status = {'status':False,'msg':'Failed to unassign, please try again!'}
                print(e)
            return HttpResponse(json.dumps(status))
              

    def post(self,request,action):
        # Check if action in defined list
        if action in ['add','clone','edit','delete']:
            # Get software list
            software_list = Software.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the software exist
        if action in ['clone','edit','delete']:
            if request.GET.get('pk'):
                software = software_list.filter(pk=request.GET.get('pk'))
                if software.exists():
                    software = software.first()
                else:
                    return HttpResponse('The software do not exist!')
            else:
                return HttpResponse('Please provide pk for the software!')        

        # Run add or clone logic
        if action in ['add','clone']:
            if action == 'add':
                title = 'Software Add'
            if action == 'clone':
                title = 'Software Clone'
            form = self.form_class(request.POST)
            if form.is_valid():
                form.instance.create_by = request.user.username
                form.save()
                return redirect(self.list_url)  
            return render(request,self.form_template,{'form':form,'title':title})

        # Run edit logic
        if action == 'edit':
            title = 'Software Edit'           
            form = self.form_class(request.POST,instance=software)
            if form.is_valid():
                if form.changed_data:       
                    form.save()
                return redirect(self.list_url)              
            return render(request,self.form_template,{'form':form,'title':title})


        # Run delete logic
        if action == 'delete':
            software_list.filter(pk=request.GET.get('pk')).delete()
            return redirect(self.list_url)              
        
# View for IP
class IpView(LoginRequiredMixin,PermissionRequiredMixin,View):
    permission_required = 'itdc.view_asset'    
    model_class = IpAddress
    form_class = IpForm
    list_template = 'itdc/ip/list.html'
    form_template = 'itdc/ip/form.html'
    detail_template = 'itdc/ip/detail.html'
    delete_template = 'itdc/ip/delete.html'
    list_url = '/itdc/ip/list'
    def get(self,request,action):
        # Check if action in defined list
        if action in ['list','detail','edit','assign','unassign','delete']:
            # Get software list
            ip_list = self.model_class.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the ip exist
        if action in ['detail','edit','delete']:
            if request.GET.get('pk'):
                ip = ip_list.filter(pk=request.GET.get('pk'))
                if ip.exists():
                    ip = ip.first()
                else:
                    return HttpResponse('The software do not exist!')
            else:
                return HttpResponse('Please provide pk for the IP!')                    
             
        # Return list page
        if action == 'list':  
            form = IpAddForm()                   
            return render(request,self.list_template,locals())
        
         # Return page for edit
        if action == 'edit':
            title = 'IP Edit'
            form = self.form_class(instance=ip)            
            return render(request,self.form_template,locals())    

        # Return detail page
        if action == 'detail':  
            available_assets = Asset.objects.filter(category__in=['pm','vm'],is_active=True).order_by('-id')          
            return render(request,self.detail_template,locals())   

        # Return delete page
        if action == 'delete':            
            return render(request,self.delete_template,{'list_url':self.list_url,'ip':ip})  

        # Run assign logic
        if action == 'assign':
            asset = Asset.objects.filter(pk=request.GET.get('assetid')).first()
            try:
                ip_list.filter(pk=request.GET.get('pk')).update(asset=asset)
                status = {'status':True,'msg':'Finish assign successfully!'}
            except Exception as e:
                status = {'status':False,'msg':'Failed to assign, please try again!'}
                print(e)
            return HttpResponse(json.dumps(status))

        # Run unassign logic
        if action == 'unassign':
            try:
                ip_list.filter(pk=request.GET.get('pk')).update(asset=None)
                status = {'status':True,'msg':'Finish unassign successfully!'}
            except Exception as e:
                status = {'status':False,'msg':'Failed to unassign, please try again!'}
                print(e)
            return HttpResponse(json.dumps(status))
              

    def post(self,request,action):
        # Check if action in defined list
        if action in ['add','edit','delete']:
            # Get IP list
            ip_list = self.model_class.objects.all().order_by('-id')  
        else:
            return HttpResponse('The action is not defined')
        
        # Check if the IP exist
        if action in ['edit','delete']:
            if request.GET.get('pk'):
                ip = ip_list.filter(pk=request.GET.get('pk'))
                if ip.exists():
                    ip = ip.first()
                else:
                    return HttpResponse('The software do not exist!')
            else:
                return HttpResponse('Please provide pk for the software!')        

        # Run add logic
        if action == 'add':
            form = IpAddForm(request.POST)
            if form.is_valid():
                ip_start = request.POST.get('ip_start')
                ip_end = request.POST.get('ip_end')
                location = request.POST.get('location')
                ip_split = ip_start.split('.')
                ip_list = []
                while int(ip_split[3]) <= int(ip_end.split('.')[3]):
                    current_ip = IpAddress(location=location,address='%s.%s.%s.%s'%(ip_split[0],ip_split[1],ip_split[2],ip_split[3]),comment=request.POST.get('comment'))
                    # Add the IP if it do not exist
                    if not IpAddress.objects.filter(address=current_ip.address):
                        ip_list.append(current_ip)
                    ip_split[3] = int(ip_split[3]) + 1
                try:
                    if ip_list:
                        IpAddress.objects.bulk_create(ip_list)
                        result = {'status':True,'msg':'IP(s) created!'}
                    else:
                        result = {'status':False,'msg':'IP(s) already exist, no need to create!'}
                except Exception as e:
                    result = {'status':False,'msg':'Failed to create!'}
                    print(e)
            else:
                result = {'status':False,'msg':'Please type in correct IP address and location!'}
                print(form.errors)
            return HttpResponse(json.dumps(result))            
            

        # Run edit logic
        if action == 'edit':
            title = 'IP Edit'           
            form = self.form_class(request.POST,instance=ip)
            if form.is_valid():
                if form.changed_data:       
                    form.save()
                return redirect(self.list_url)              
            return render(request,self.form_template,{'form':form,'title':title})


        # Run delete logic
        if action == 'delete':
            ip_list.filter(pk=request.GET.get('pk')).delete()
            return redirect(self.list_url)              
        
# View for report
class ReportView(LoginRequiredMixin,View):
    def get(self,request,reportname):
        if reportname not in ['diagram']:
            return HttpResponse('The report name do not exist!')
        if reportname == 'diagram':
            if request.GET.get('location') not in ['all','beijing','shanghai','kunshan','singapore','xiaoshan','india','dia']:
                return HttpResponse('The location do not exist!')
            draw_diagram(request.GET.get('location'))
            return render(request,'itdc/report/diagram.html',{'location':request.GET.get('location')})
            
# View to import asset
class ImportView(LoginRequiredMixin,View):     

    def get(self,request,tablename):
        # Check if the table is defined
        if tablename not in ['asset',]:
            return HttpResponse('The table name is not defined!')
        if tablename == 'asset':
            title = 'Asset Import'
            return render(request,'itdc/import/data_import.html',locals())
        pass
    def post(self,request,tablename):
        # Check if the table is defined
        if tablename not in ['asset',]:
            return HttpResponse('The table name is not defined!')
    
        excel_file = request.FILES['csv_file']
        # Get file data
        wb = load_workbook(excel_file,read_only=True,data_only=True)
        sheet_obj = wb[wb.sheetnames[0]]
        # Get table head
        current_row = 0
        result_list = []
        for row in sheet_obj.iter_rows(values_only=True):
            if current_row == 0:
                current_row += 1
                tableh = list(row)
                continue
            else:
                row = list(row)
                # Convert data to dictionary
                data_dic = {}
                i = 0
                for column_name in tableh:
                    data_dic[column_name] = row[i]
                    i += 1
            current_row += 1

            # Import assets:
            if tablename == 'asset':
                title = 'Result for asset import'
                try:
                    new_asset = Asset(**data_dic)
                    new_asset.create_by = request.user.username
                    if data_dic.get('create_by','') != '':
                        new_asset.create_by = data_dic.get('create_by')                                      
                    new_asset.save()  
                    if data_dic.get('create_time','') != '':
                        new_asset.create_time = data_dic.get('create_time')  
                        new_asset.save()                                                        
                    result = 'Success'
                except Exception as e:
                    result =  str(e)

            result_list.append([result] + row)    
        tableh = ['Result'] + tableh                   
        return render(request,'itdc/import/import_result.html',locals())           